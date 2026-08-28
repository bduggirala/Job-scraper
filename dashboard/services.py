"""Non-UI half of the dashboard: run control, run state, workbook editing.

Deliberately Streamlit-free. Everything here is a plain function over the
files the scraper already writes (``output/last_run.json``,
``output/company_jobs.*``, ``logs/scraper.log``) plus two small files the
dashboard owns:

``output/dashboard_run.lock``
    Exists only while a dashboard-launched run is in flight. Created with
    ``O_EXCL`` so the check-and-claim is atomic across processes, not just
    across browser tabs, and holds the supervisor's PID so a crashed run is
    recognisable rather than permanent.
``output/dashboard_last_run.json``
    The most recent launch's *outcome* - start, finish, and the scraper's real
    exit code. Overwritten each run: this is a status file, not a history.

The scraper itself is never reimplemented. A run is ``python main.py`` in a
child process, exactly as it is from a terminal.
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence
from urllib.parse import urlsplit, urlunsplit

import pandas as pd
from openpyxl import load_workbook

# Reused rather than restated: the dashboard must escape exactly the values the
# pipeline's own CSV guard escapes, and two copies of that tuple would drift.
from pipeline import (
    _FORMULA_PREFIXES,
    load_companies,
    resolve_companies_path,
    retryable_from_report,
)
from settings import PROJECT_ROOT, Settings, load_settings

MAIN_PY = PROJECT_ROOT / "main.py"

#: Files the dashboard owns. Single current files, never timestamped, never
#: accumulated - the same policy ``logs/scraper.log`` follows.
RUN_LOCK_NAME = "dashboard_run.lock"
RUN_STATE_NAME = "dashboard_last_run.json"
RUN_LOG_NAME = "dashboard_run.log"
WORKBOOK_LOCK_SUFFIX = ".dashboard.lock"
#: At most one recoverable backup, overwritten each write. The marker goes
#: before the extension, not after it, so the backup is a file Excel and
#: openpyxl can open directly - a recovery copy nothing can read is not one.
WORKBOOK_BACKUP_MARKER = ".bak-dashboard"

#: A lock whose PID is not yet known is only trusted this long - it covers the
#: gap between claiming the lock and the supervisor being spawned.
STARTUP_GRACE_SECONDS = 60.0
#: Absolute ceiling on a run, well past the summed phase budgets in
#: ``config/settings.yaml`` (30 min API + 50 min browser). Guards against a PID
#: that was recycled by the OS onto an unrelated live process.
MAX_RUN_SECONDS = 6 * 60 * 60
#: How long a workbook write waits for a competing writer before giving up.
WORKBOOK_LOCK_TIMEOUT_SECONDS = 10.0
WORKBOOK_LOCK_STALE_SECONDS = 300.0

#: Run outcomes the dashboard reports. ``partial`` is the scraper's own word
#: for "the rows are real, the coverage is not".
STATUS_IDLE = "idle"
STATUS_RUNNING = "running"
STATUS_COMPLETED = "completed"
STATUS_PARTIAL = "partial"
STATUS_FAILED = "failed"
STATUS_STALE = "stale"

#: Per-company outcomes that belong in the "needs attention" table. Not every
#: one of them is retryable: ``blocked`` means the site issued a challenge, and
#: ``--retry-failed`` deliberately leaves those alone (see
#: ``pipeline.RETRYABLE_STATUSES``).
PROBLEM_STATUSES = ("partial", "failed", "blocked")

_RUN_ID_RE = re.compile(r"^(\d{8})T(\d{6})Z$")

#: Progress is read out of the log the run already writes, rather than by
#: instrumenting the pipeline. ``ats.router`` logs one line as each company is
#: routed for collection, and ``pipeline`` logs the total up front, so the two
#: together give a real "23 of 183" without the scraper having to change.
_ROUTER_LINE = re.compile(r"\bats\.router: (?P<company>.+?) -> (?P<detail>\S.*?)\s*$")
_EXECUTING_LINE = re.compile(r"\bpipeline: Executing (?P<total>\d+) companies\b")
#: ``ats.router`` logs a company twice - once with the provider it routed to,
#: once with the job count when it finishes. Telling the two apart is what
#: makes "finished" a real number rather than a count of log lines.
_RETRIEVED = re.compile(r"^\d+ jobs? retrieved\b")


class DashboardError(RuntimeError):
    """A user-facing problem: the message is meant to be shown verbatim."""


class RunAlreadyActive(DashboardError):
    """A scraper run is already in flight; a second one must not start."""


class WorkbookBusy(DashboardError):
    """The workbook cannot be written right now (Excel, or a run, has it)."""


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

def output_dir(settings: Settings | None = None) -> Path:
    cfg = settings or load_settings()
    return cfg.resolve_path("output.directory", "output")


def log_path(settings: Settings | None = None) -> Path:
    """The scraper's own single current log."""
    cfg = settings or load_settings()
    return cfg.resolve_path("logging.file", "logs/scraper.log")


def run_log_path(settings: Settings | None = None) -> Path:
    """The console transcript of the current dashboard-launched run.

    One file, truncated at the start of each run - the same policy
    ``logs/scraper.log``, ``logs/canary.log`` and ``logs/discovery.log``
    follow. It exists because a scraper that dies before it can log (a bad
    config, an import error) still prints why, and that is exactly the run the
    dashboard has to explain.
    """
    return log_path(settings).with_name(RUN_LOG_NAME)


def run_lock_path(settings: Settings | None = None) -> Path:
    return output_dir(settings) / RUN_LOCK_NAME


def run_state_path(settings: Settings | None = None) -> Path:
    return output_dir(settings) / RUN_STATE_NAME


def last_run_report_path(settings: Settings | None = None) -> Path:
    return output_dir(settings) / "last_run.json"


def output_files(settings: Settings | None = None) -> dict[str, Path]:
    """The current full-run outputs, by kind. Paths may not exist yet."""
    cfg = settings or load_settings()
    out = output_dir(cfg)
    return {
        "csv": out / cfg.get("output.csv", "company_jobs.csv"),
        "xlsx": out / cfg.get("output.xlsx", "company_jobs.xlsx"),
        "json": out / cfg.get("output.json", "company_jobs.json"),
        "failures": out / cfg.get("output.failures", "scraper_failures.csv"),
    }


# ---------------------------------------------------------------------------
# Time helpers. Everything is stored and compared in UTC; only the display
# layer ever formats.
# ---------------------------------------------------------------------------

def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def parse_timestamp(value: Any) -> datetime | None:
    """Parse an ISO-8601 timestamp into an aware UTC datetime, or ``None``.

    Tolerant on purpose: this reads files a crashed run may have left half
    written, and a bad timestamp must not take the page down with it.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        moment = value
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            moment = datetime.fromisoformat(text)
        except ValueError:
            return None
    return moment.astimezone(timezone.utc) if moment.tzinfo else moment.replace(tzinfo=timezone.utc)


def parse_run_id(run_id: Any) -> datetime | None:
    """``20260827T202307Z`` -> the moment that run started, in UTC.

    The run id *is* the start time - the scraper stamps it at the top of
    ``run()`` - so a report on disk carries its own start even for a run this
    dashboard never launched.
    """
    if not run_id:
        return None
    match = _RUN_ID_RE.match(str(run_id).strip())
    if not match:
        return None
    try:
        return datetime.strptime(
            f"{match.group(1)}{match.group(2)}", "%Y%m%d%H%M%S"
        ).replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def humanize_duration(seconds: float | None) -> str:
    """``"1h 04m"`` / ``"12 minutes"`` / ``"45 seconds"``."""
    if seconds is None:
        return ""
    seconds = max(0.0, float(seconds))
    if seconds < 60:
        whole = int(seconds)
        return f"{whole} second{'' if whole == 1 else 's'}"
    if seconds < 3600:
        minutes = int(seconds // 60)
        return f"{minutes} minute{'' if minutes == 1 else 's'}"
    hours, remainder = divmod(int(seconds), 3600)
    return f"{hours}h {remainder // 60:02d}m"


def humanize_age(moment: datetime | None, *, now: datetime | None = None) -> str:
    """``"12 minutes ago"``. Empty string when there is no timestamp."""
    if moment is None:
        return ""
    seconds = ((now or utcnow()) - moment).total_seconds()
    if seconds < 0:
        return "just now"
    return f"{humanize_duration(seconds)} ago"


def format_utc(moment: datetime | None) -> str:
    return moment.strftime("%Y-%m-%d %H:%M:%S UTC") if moment else "-"


def format_local(moment: datetime | None) -> str:
    """The same instant in this machine's timezone.

    The project configures no display timezone, so the dashboard shows UTC as
    the authoritative value and the local rendering beside it rather than
    silently picking one.
    """
    if moment is None:
        return "-"
    return moment.astimezone().strftime("%Y-%m-%d %H:%M:%S %Z").strip()


def format_clock(moment: datetime | None, *, seconds: bool = True) -> str:
    """Local time as a person reads it: ``Aug 27, 2026 - 4:23:07 PM EDT``.

    12-hour with AM/PM, a spelled-out month and no leading zeros, because the
    ISO form this replaced (``2026-08-27 20:23:07 UTC``) ran together into an
    unreadable digit string at a glance. UTC is still shown alongside wherever
    this is used - it stays the authoritative value, this is the legible one.

    ``%Z`` is spelled out on Windows ("Eastern Daylight Time"), which is far
    too long for a column heading, so a multi-word zone is initialised to its
    usual abbreviation.
    """
    if moment is None:
        return "-"
    local = moment.astimezone()
    day = f"{local.strftime('%b')} {local.day}, {local.year}"
    clock = local.strftime("%I:%M:%S %p" if seconds else "%I:%M %p").lstrip("0")
    zone = local.strftime("%Z")
    if len(zone) > 5:
        zone = "".join(word[0] for word in zone.split() if word[:1].isupper())
    return f"{day} - {clock}" + (f" {zone}" if zone else "")


# ---------------------------------------------------------------------------
# Process liveness
# ---------------------------------------------------------------------------

def pid_alive(pid: int | None) -> bool:
    """Is this PID a live process?

    ``os.kill(pid, 0)`` is the usual answer and is **wrong on Windows**:
    CPython implements ``os.kill`` there as ``TerminateProcess(handle, sig)``,
    so probing with signal 0 would kill the running scraper with exit code 0 -
    and the dashboard would then report that killed run as a clean success.
    Windows therefore goes through ``OpenProcess``/``GetExitCodeProcess``.
    """
    if not pid or int(pid) <= 0:
        return False
    if os.name == "nt":
        import ctypes
        from ctypes import wintypes

        PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
        ERROR_ACCESS_DENIED = 5
        STILL_ACTIVE = 259

        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        kernel32.OpenProcess.restype = wintypes.HANDLE
        handle = kernel32.OpenProcess(PROCESS_QUERY_LIMITED_INFORMATION, False, int(pid))
        if not handle:
            # Access denied means the process exists but belongs to someone
            # else; anything else means it is gone. Erring toward "alive" here
            # is the safe direction - it blocks a second run rather than
            # allowing two.
            return ctypes.get_last_error() == ERROR_ACCESS_DENIED
        try:
            code = wintypes.DWORD()
            if kernel32.GetExitCodeProcess(handle, ctypes.byref(code)):
                return code.value == STILL_ACTIVE
            return True
        finally:
            kernel32.CloseHandle(handle)
    try:
        os.kill(int(pid), 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True
    except (OSError, ValueError):
        return False
    return True


# ---------------------------------------------------------------------------
# The run lock
# ---------------------------------------------------------------------------

def _read_json(path: Path) -> dict | None:
    """Read a JSON object, or ``None`` if it is missing, empty or malformed."""
    try:
        text = Path(path).read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError):
        return None
    if not text.strip():
        return None
    try:
        data = json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return None
    return data if isinstance(data, dict) else None


def read_run_lock(settings: Settings | None = None) -> dict | None:
    """The current lock's contents, or ``None`` when no lock file exists.

    A lock that exists but cannot be parsed still counts as a lock - it is
    reported with an unknown PID, which the staleness rule then times out.
    """
    path = run_lock_path(settings)
    if not path.exists():
        return None
    return _read_json(path) or {"pid": None, "started_at": None, "unreadable": True}


def lock_state(
    settings: Settings | None = None, *, now: datetime | None = None
) -> tuple[str, dict | None]:
    """``("idle" | "running" | "stale", lock_contents_or_None)``.

    ``stale`` means a lock file survived the process that owned it - a crash, a
    kill, or a machine that went down mid-run. It is never cleared silently;
    :func:`clear_stale_lock` is an explicit action.
    """
    lock = read_run_lock(settings)
    if lock is None:
        return STATUS_IDLE, None

    moment = now or utcnow()
    started = parse_timestamp(lock.get("started_at"))
    age = (moment - started).total_seconds() if started else None

    if age is not None and age > MAX_RUN_SECONDS:
        return STATUS_STALE, lock

    pid = lock.get("pid")
    if pid is None:
        # Claimed but not yet handed to a supervisor. Real for a moment,
        # abandoned if it stays that way.
        if age is None or age > STARTUP_GRACE_SECONDS:
            return STATUS_STALE, lock
        return STATUS_RUNNING, lock

    return (STATUS_RUNNING, lock) if pid_alive(pid) else (STATUS_STALE, lock)


def is_run_active(settings: Settings | None = None) -> bool:
    """True only while a live supervisor holds the lock.

    This is the single gate for both "don't start another run" and "don't
    touch the workbook".
    """
    return lock_state(settings)[0] == STATUS_RUNNING


def clear_stale_lock(settings: Settings | None = None) -> bool:
    """Remove a lock whose owner is gone. Refuses while the owner is alive."""
    state, lock = lock_state(settings)
    if state == STATUS_IDLE:
        return False
    if state == STATUS_RUNNING:
        raise RunAlreadyActive(
            f"The run holding this lock is still alive (PID {(lock or {}).get('pid')}). "
            "Wait for it to finish, or stop that process first."
        )
    try:
        run_lock_path(settings).unlink()
    except FileNotFoundError:
        return False
    return True


def _write_lock(path: Path, payload: dict) -> None:
    Path(path).write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def _child_environment() -> dict[str, str]:
    """Environment for the scraper child.

    ``EMAIL_ENABLED`` is forced off. ``settings.load_env_file`` never overrides
    a variable already present in the environment, so setting it here means a
    ``.env`` on this machine cannot turn real delivery on for a run the
    dashboard started - and ``--no-email`` is passed as well. Nothing is added,
    logged or displayed: the child simply inherits, with that one override.
    """
    env = dict(os.environ)
    env["EMAIL_ENABLED"] = "false"
    env["SCRAPER_SMTP_DRY_RUN"] = "1"
    env.setdefault("PYTHONIOENCODING", "utf-8")
    return env


def scraper_argv(*, dry_run: bool = False, extra: Sequence[str] = ()) -> list[str]:
    """The ``main.py`` flags a dashboard run uses.

    ``--no-email`` is not optional: the dashboard never sends a real digest.
    """
    argv = ["--no-email"]
    if dry_run:
        argv.append("--dry-run")
    argv.extend(extra)
    return argv


def start_run(
    settings: Settings | None = None,
    *,
    dry_run: bool = False,
    extra_args: Sequence[str] = (),
    popen: Any = None,
) -> dict:
    """Launch the real scraper in a supervised child process.

    The lock is claimed with ``O_CREAT | O_EXCL`` *before* anything is spawned,
    so two dashboards (or two browser tabs) racing the button cannot both win:
    the loser gets :class:`RunAlreadyActive`.

    ``popen`` is an injection point for tests - the production call is a plain
    argument-list ``subprocess.Popen`` with ``shell=False``.

    Returns the lock payload that was written.
    """
    cfg = settings or load_settings()
    out = output_dir(cfg)
    out.mkdir(parents=True, exist_ok=True)
    lock_file = run_lock_path(cfg)

    state, existing = lock_state(cfg)
    if state == STATUS_RUNNING:
        raise RunAlreadyActive(
            f"A scraper run is already in progress (PID {(existing or {}).get('pid')}, started "
            f"{(existing or {}).get('started_at')}). Only one run may be active at a time."
        )
    if state == STATUS_STALE:
        raise RunAlreadyActive(
            "A previous run left a stale lock behind. Clear it before starting a new run."
        )

    argv = scraper_argv(dry_run=dry_run, extra=extra_args)
    started_at = utcnow()
    transcript = run_log_path(cfg)
    transcript.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "pid": None,
        "started_at": started_at.isoformat(),
        "args": argv,
        "dry_run": bool(dry_run),
        "log": str(transcript),
        "scraper_log": str(log_path(cfg)),
        "state_file": str(run_state_path(cfg)),
    }

    try:
        handle = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:  # lost the race between the check and here
        raise RunAlreadyActive(
            "A scraper run was started by someone else a moment ago."
        ) from exc
    os.close(handle)
    _write_lock(lock_file, payload)

    command = [
        sys.executable, "-m", "dashboard.runner",
        "--lock", str(lock_file),
        "--state", str(run_state_path(cfg)),
        "--log", str(transcript),
        "--started-at", started_at.isoformat(),
        "--", *argv,
    ]

    spawn = popen or subprocess.Popen
    kwargs: dict[str, Any] = {
        "cwd": str(PROJECT_ROOT),
        "env": _child_environment(),
        "stdin": subprocess.DEVNULL,
        "stdout": subprocess.DEVNULL,
        "stderr": subprocess.DEVNULL,
        "close_fds": True,
    }
    if os.name == "nt":
        # Ctrl-C in the terminal running Streamlit sends a console event to the
        # whole group; a scraper 30 minutes into a run should not die with it.
        kwargs["creationflags"] = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)

    try:
        # shell=False (the default): an argument list, never an interpolated
        # command string, so nothing here can be shell-injected.
        process = spawn(command, **kwargs)
    except Exception as exc:
        lock_file.unlink(missing_ok=True)
        raise DashboardError(f"Could not start the scraper: {exc}") from exc

    payload["pid"] = getattr(process, "pid", None)
    _write_lock(lock_file, payload)
    return payload


# ---------------------------------------------------------------------------
# Reading what a run left behind
# ---------------------------------------------------------------------------

def load_run_state(settings: Settings | None = None) -> dict | None:
    """The dashboard's record of the last launch: exit code, start, finish."""
    return _read_json(run_state_path(settings))


def load_last_run(settings: Settings | None = None) -> dict | None:
    """``output/last_run.json``, or ``None`` if absent or malformed.

    Missing and malformed are deliberately the same answer *here* and
    different in the UI, which also reports whether the file exists - "no run
    has happened yet" and "the report is corrupt" call for different actions.
    """
    data = _read_json(last_run_report_path(settings))
    if data is None or "companies" not in data:
        return None
    return data


def read_log_tail(path: Path, lines: int = 200) -> str:
    """The last ``lines`` lines of a log, or an explanatory string."""
    path = Path(path)
    try:
        if not path.exists():
            return ""
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            return "".join(handle.readlines()[-lines:])
    except OSError as exc:
        return f"(could not read {path.name}: {exc})"


@dataclass
class RunProgress:
    """How far into the company list the current run has got."""

    current_company: str = ""
    provider: str = ""
    started: int = 0
    finished: int = 0
    total: int | None = None

    @property
    def fraction(self) -> float | None:
        """Companies finished, out of the total the run announced."""
        if not self.total:
            return None
        return min(1.0, self.finished / self.total)


def parse_progress(log_text: str) -> RunProgress:
    """Read progress out of the scraper's own log text.

    Deliberately a pure function over the text: the log is the only progress
    signal that exists mid-run, and parsing it here keeps that parsing
    testable without a running scraper.
    """
    progress = RunProgress()
    started: dict[str, str] = {}
    finished: set[str] = set()
    for line in log_text.splitlines():
        total = _EXECUTING_LINE.search(line)
        if total:
            progress.total = int(total.group("total"))
            # Routing is announced once; anything logged before it belongs to
            # the resolution phase and is not a company being collected.
            started.clear()
            finished.clear()
            continue
        routed = _ROUTER_LINE.search(line)
        if not routed:
            continue
        company = routed.group("company").strip()
        detail = routed.group("detail").strip()
        progress.current_company = company
        if _RETRIEVED.match(detail):
            finished.add(company)
            # A completion line carries a job count, not a provider - keep the
            # provider this company was routed to instead of overwriting it.
            progress.provider = started.get(company, "")
        else:
            started.setdefault(company, detail)
            progress.provider = detail
    progress.started = len(started)
    progress.finished = len(finished)
    return progress


def run_progress(settings: Settings | None = None) -> RunProgress:
    """Progress of the run currently writing ``logs/scraper.log``.

    The whole file is read rather than a tail: each run truncates the log, so
    it only ever holds this run, and the started-company count needs all of it.
    """
    path = log_path(settings)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return RunProgress()
    return parse_progress(text)


def file_age(path: Path, *, now: datetime | None = None) -> tuple[datetime | None, str]:
    """``(modified_utc, "12 minutes ago")`` for an output file."""
    try:
        modified = datetime.fromtimestamp(Path(path).stat().st_mtime, tz=timezone.utc)
    except OSError:
        return None, ""
    return modified, humanize_age(modified, now=now)


@dataclass
class RunStatus:
    """Everything the Run Scraper tab shows, resolved from files on disk."""

    status: str = STATUS_IDLE
    detail: str = ""
    running: bool = False
    stale_lock: bool = False
    lock: dict | None = None

    started_at: datetime | None = None
    finished_at: datetime | None = None
    duration_seconds: float | None = None
    exit_code: int | None = None
    dry_run: bool = False

    run_id: str = ""
    report_generated_at: datetime | None = None
    companies_attempted: int | None = None
    status_counts: dict[str, int] = field(default_factory=dict)
    totals: dict[str, Any] = field(default_factory=dict)
    report_available: bool = False
    report_malformed: bool = False

    error_message: str = ""

    @property
    def successful_companies(self) -> int:
        return int(self.status_counts.get("success", 0))

    @property
    def partial_companies(self) -> int:
        return int(self.status_counts.get("partial", 0))

    @property
    def failed_companies(self) -> int:
        return int(self.status_counts.get("failed", 0))

    @property
    def blocked_companies(self) -> int:
        return int(self.status_counts.get("blocked", 0))

    @property
    def no_jobs_companies(self) -> int:
        return int(self.status_counts.get("no_jobs", 0))


def _verdict_from_report(report: dict) -> str:
    """``completed`` / ``partial`` / ``failed`` from a report's status counts."""
    counts = report.get("status_counts") or {}
    troubled = sum(int(counts.get(key, 0)) for key in PROBLEM_STATUSES)
    reached = sum(int(counts.get(key, 0)) for key in ("success", "no_jobs"))
    if troubled == 0:
        return STATUS_COMPLETED if reached else STATUS_FAILED
    return STATUS_PARTIAL if reached else STATUS_FAILED


def run_status(
    settings: Settings | None = None, *, now: datetime | None = None
) -> RunStatus:
    """Assemble the current run picture from lock + state file + run report.

    The order matters. A live lock means *running*, whatever the files say. An
    exit code decides success or failure - never the fact that a process was
    launched. Only then does the run report refine ``completed`` into
    ``partial``.
    """
    cfg = settings or load_settings()
    moment = now or utcnow()
    status = RunStatus()

    state, lock = lock_state(cfg, now=moment)
    status.lock = lock

    report_path = last_run_report_path(cfg)
    report = load_last_run(cfg)
    status.report_available = report is not None
    status.report_malformed = report_path.exists() and report is None

    if report:
        status.run_id = str(report.get("run_id") or "")
        status.report_generated_at = parse_timestamp(report.get("generated_at"))
        status.companies_attempted = report.get("companies_attempted")
        status.status_counts = {
            str(key): int(value)
            for key, value in (report.get("status_counts") or {}).items()
        }
        status.totals = report.get("totals") or {}

    if state == STATUS_RUNNING:
        status.status = STATUS_RUNNING
        status.running = True
        status.started_at = parse_timestamp((lock or {}).get("started_at"))
        status.dry_run = bool((lock or {}).get("dry_run"))
        if status.started_at:
            status.duration_seconds = (moment - status.started_at).total_seconds()
        status.detail = "The scraper is running. The workbook is read-only until it finishes."
        return status

    if state == STATUS_STALE:
        status.stale_lock = True

    launch = load_run_state(cfg)
    if launch:
        status.started_at = parse_timestamp(launch.get("started_at"))
        status.finished_at = parse_timestamp(launch.get("finished_at"))
        status.exit_code = launch.get("exit_code")
        status.dry_run = bool(launch.get("dry_run"))
        status.error_message = str(launch.get("error") or "")

    # A run report on disk may predate this dashboard entirely; fall back to
    # the run id, which is the start moment the scraper stamped itself.
    if status.started_at is None:
        status.started_at = parse_run_id(status.run_id)
    if status.finished_at is None and status.report_generated_at:
        status.finished_at = status.report_generated_at
    if status.duration_seconds is None and status.started_at and status.finished_at:
        status.duration_seconds = (status.finished_at - status.started_at).total_seconds()

    if status.stale_lock:
        status.status = STATUS_FAILED
        status.detail = (
            "A previous run left a lock behind without finishing - the process is gone. "
            "Clear the stale lock to run again."
        )
        return status

    if status.exit_code is None:
        if report:
            status.status = _verdict_from_report(report)
            status.detail = "Showing the last run recorded in output/last_run.json."
        else:
            status.status = STATUS_IDLE
            status.detail = (
                "output/last_run.json is unreadable."
                if status.report_malformed
                else "No run has been recorded yet."
            )
        return status

    if int(status.exit_code) != 0:
        status.status = STATUS_FAILED
        status.detail = f"The scraper exited with code {status.exit_code}."
        return status

    if status.dry_run:
        status.status = STATUS_COMPLETED
        status.detail = (
            "Dry run finished: routing only, no jobs scraped and no outputs rewritten."
        )
        return status

    status.status = _verdict_from_report(report) if report else STATUS_COMPLETED
    status.detail = (
        "Run finished cleanly."
        if status.status == STATUS_COMPLETED
        else "Run finished, but some companies did not complete."
    )
    return status


def load_current_jobs(settings: Settings | None = None) -> pd.DataFrame:
    """The validated current job export, straight from ``company_jobs.csv``.

    The scraper already produces this file; the dashboard reads it rather than
    re-deriving the rows, so what is displayed is exactly what was exported.
    """
    path = output_files(settings)["csv"]
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    except (OSError, UnicodeDecodeError, pd.errors.ParserError, pd.errors.EmptyDataError):
        return pd.DataFrame()


def change_status_counts(jobs: pd.DataFrame) -> dict[str, int]:
    """``{new, changed, unchanged}`` from the export's own ``change_status``."""
    if jobs.empty or "change_status" not in jobs.columns:
        return {}
    counts = jobs["change_status"].astype(str).str.strip().str.lower().value_counts()
    return {str(key): int(value) for key, value in counts.items() if key and key != "nan"}


def within_window_count(jobs: pd.DataFrame) -> int:
    """Rows the freshness filter placed inside the configured window."""
    if jobs.empty or "date_filter_status" not in jobs.columns:
        return 0
    return int((jobs["date_filter_status"].astype(str).str.strip() == "within_window").sum())


def problem_companies(report: dict | None) -> pd.DataFrame:
    """Partial / failed / blocked companies with why they stopped."""
    if not report:
        return pd.DataFrame()
    rows = [
        {
            "Company": row.get("company"),
            "Status": row.get("status"),
            "Provider": row.get("provider"),
            "Method": row.get("method"),
            "Jobs": row.get("jobs"),
            "Reported total": row.get("reported_total"),
            "Stop reason": row.get("stop_reason"),
            "Error": row.get("error_type"),
            "Detail": row.get("error_message"),
            "Removal sync": "allowed" if row.get("removal_sync_allowed") else "withheld",
        }
        for row in (report.get("companies") or [])
        if row.get("status") in PROBLEM_STATUSES
    ]
    frame = pd.DataFrame(rows)
    if not frame.empty:
        order = {status: index for index, status in enumerate(PROBLEM_STATUSES)}
        frame = frame.sort_values(
            "Status", key=lambda column: column.map(order), kind="stable"
        ).reset_index(drop=True)
    return frame


def retry_targets(report: dict | None) -> list[str]:
    """The companies a ``--retry-failed`` run would actually re-attempt.

    A subset of :func:`problem_companies`: the retry list is the pipeline's,
    not the dashboard's, so a blocked company shown in the attention table is
    correctly absent here. Returns an empty list rather than raising - the
    caller is drawing a button, and "no report yet" and "nothing to retry" both
    mean the same thing to a button: it is disabled.
    """
    return retryable_from_report(report)


@dataclass
class RetryRecord:
    """What the last merged retry did, from the report it merged itself into."""

    run_id: str = ""
    finished_at: datetime | None = None
    companies: list[str] = field(default_factory=list)


def last_retry(report: dict | None) -> RetryRecord | None:
    """The ``last_retry`` block a merged retry leaves in ``last_run.json``.

    Without it the page would show a retry's numbers with nothing saying a
    retry produced them - the cost of merging into one report rather than
    writing a second one beside it.
    """
    block = (report or {}).get("last_retry") or {}
    run_id = str(block.get("run_id") or "")
    if not run_id:
        return None
    return RetryRecord(
        run_id=run_id,
        finished_at=parse_timestamp(block.get("finished_at")),
        companies=[str(name) for name in (block.get("companies") or [])],
    )


# ---------------------------------------------------------------------------
# Workbook editing
# ---------------------------------------------------------------------------

def workbook_columns(settings: Settings | None = None) -> dict[str, str]:
    """The three columns the loader requires, under their workbook headings."""
    cfg = settings or load_settings()
    columns = cfg.get("columns", {}) or {}
    return {
        "company": columns.get("company", "Company"),
        "ats_url": columns.get("ats_url", "ATS URL"),
        "live_jobs_url": columns.get(
            "live_jobs_url", "Live Jobs Page (if ATS URL unavailable)"
        ),
    }


def workbook_path(settings: Settings | None = None) -> Path:
    return resolve_companies_path(settings or load_settings())


def read_workbook(settings: Settings | None = None) -> pd.DataFrame:
    """Every column of the active sheet, exactly as stored.

    Deliberately not :func:`pipeline.load_companies` - that returns the three
    columns the router needs, and the Manage Companies tab has to show the
    workbook's real schema, including ``Data Retrieved``, ``Jobs Found`` and
    the discovery columns.
    """
    cfg = settings or load_settings()
    path = workbook_path(cfg)
    if not path.exists():
        raise DashboardError(f"Workbook not found: {path}")
    sheet = cfg.get("input_sheet")
    return pd.read_excel(path, sheet_name=sheet if sheet else 0)


def workbook_backup_path(path: Path) -> Path:
    """The single recoverable copy kept beside the workbook."""
    path = Path(path)
    return path.with_name(f"{path.stem}{WORKBOOK_BACKUP_MARKER}{path.suffix}")


def excel_lock_file(path: Path) -> Path:
    """Excel's owner file, e.g. ``~$companies.xlsx`` beside the workbook."""
    path = Path(path)
    return path.with_name(f"~${path.name}")


def workbook_is_open_in_excel(path: Path) -> bool:
    """Is the workbook currently held open by Excel?

    Two signals, because either alone misses cases: Excel's ``~$`` owner file,
    and an actual attempt to open the file for writing (Windows refuses while
    another process holds a write share).
    """
    path = Path(path)
    if excel_lock_file(path).exists():
        return True
    if not path.exists():
        return False
    try:
        with path.open("r+b"):
            return False
    except PermissionError:
        return True
    except OSError:
        return False


class _WorkbookLock:
    """Cross-process advisory lock on the workbook, as a context manager.

    ``O_CREAT | O_EXCL`` makes the claim atomic; a lock older than
    ``WORKBOOK_LOCK_STALE_SECONDS`` is broken, because the only writes it ever
    guards are sub-second and a survivor means the writer died.
    """

    def __init__(self, path: Path, timeout: float = WORKBOOK_LOCK_TIMEOUT_SECONDS):
        self.target = Path(path)
        self.lock_path = self.target.with_name(self.target.name + WORKBOOK_LOCK_SUFFIX)
        self.timeout = timeout
        self._held = False

    def __enter__(self) -> "_WorkbookLock":
        deadline = time.monotonic() + self.timeout
        while True:
            try:
                handle = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                if self._break_if_stale():
                    continue
                if time.monotonic() >= deadline:
                    raise WorkbookBusy(
                        "Another workbook write is in progress. Try again in a moment."
                    ) from None
                time.sleep(0.1)
                continue
            with os.fdopen(handle, "w", encoding="utf-8") as fh:
                json.dump({"pid": os.getpid(), "at": utcnow().isoformat()}, fh)
            self._held = True
            return self

    def _break_if_stale(self) -> bool:
        try:
            age = time.time() - self.lock_path.stat().st_mtime
        except OSError:
            return True  # vanished between the two calls - retry immediately
        if age > WORKBOOK_LOCK_STALE_SECONDS:
            self.lock_path.unlink(missing_ok=True)
            return True
        return False

    def __exit__(self, *exc_info) -> None:
        if self._held:
            self.lock_path.unlink(missing_ok=True)
            self._held = False


def escape_formula(value: str) -> str:
    """Neutralise a spreadsheet formula in text destined for a cell.

    Same guard, and the same prefix set, the pipeline applies to scraped text
    before writing the CSV: a value beginning ``=``, ``+``, ``-``, ``@``, a tab
    or a carriage return is executed on open, so a company name typed into this
    form would otherwise be code running on whoever opens the workbook.
    """
    text = str(value)
    return f"'{text}" if text.startswith(_FORMULA_PREFIXES) else text


def normalize_name(value: Any) -> str:
    """Company name, normalised the way :func:`pipeline.load_companies` does."""
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def normalize_url(value: Any) -> str:
    """A URL reduced to what makes two entries the *same* entry.

    Case, a ``www.`` prefix, a trailing slash and a fragment are all noise for
    duplicate detection - ``https://Boards.Greenhouse.io/acme/`` and
    ``https://boards.greenhouse.io/acme`` are one ATS, not two.
    """
    text = str(value or "").strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""
    if "://" not in text:
        text = "https://" + text
    parts = urlsplit(text)
    host = parts.netloc.lower().removeprefix("www.")
    return urlunsplit(("https", host, parts.path.rstrip("/"), parts.query, ""))


def validate_url(value: str) -> str | None:
    """``None`` when the URL is usable, else why it is not."""
    text = str(value or "").strip()
    if not text:
        return None
    if any(char.isspace() for char in text):
        return f"{text!r} contains whitespace."
    parts = urlsplit(text if "://" in text else "https://" + text)
    if parts.scheme not in ("http", "https"):
        return f"{text!r} must be an http:// or https:// URL."
    if not parts.netloc or "." not in parts.netloc:
        return f"{text!r} does not look like a web address."
    return None


@dataclass
class DuplicateCheck:
    """What an existing workbook already says about a proposed company."""

    name_conflict: str = ""
    url_conflicts: list[tuple[str, str, str]] = field(default_factory=list)

    @property
    def has_name_conflict(self) -> bool:
        return bool(self.name_conflict)

    @property
    def has_url_conflict(self) -> bool:
        return bool(self.url_conflicts)


def check_duplicates(
    frame: pd.DataFrame,
    columns: dict[str, str],
    name: str,
    ats_url: str = "",
    live_url: str = "",
    *,
    ignore_row_name: str = "",
) -> DuplicateCheck:
    """Case-insensitive name match, normalised URL match, across both columns."""
    check = DuplicateCheck()
    if frame.empty:
        return check

    ignore = normalize_name(ignore_row_name).lower()
    wanted_name = normalize_name(name).lower()
    proposed = {
        columns["ats_url"]: normalize_url(ats_url),
        columns["live_jobs_url"]: normalize_url(live_url),
    }

    for _, row in frame.iterrows():
        existing_name = normalize_name(row.get(columns["company"]))
        if not existing_name:
            continue
        lowered = existing_name.lower()
        if ignore and lowered == ignore:
            continue
        if wanted_name and lowered == wanted_name:
            check.name_conflict = existing_name
        for column in (columns["ats_url"], columns["live_jobs_url"]):
            existing_url = normalize_url(row.get(column))
            if not existing_url:
                continue
            for proposed_column, proposed_url in proposed.items():
                if proposed_url and proposed_url == existing_url:
                    check.url_conflicts.append((existing_name, column, proposed_column))
    return check


def _validate_written_workbook(
    candidate: Path, settings: Settings, expected_headers: list, expected_rows: int
) -> None:
    """Prove a candidate workbook is sound before it replaces the real one.

    Three checks, in increasing strength: it opens, its header row and row
    count are what we meant to write, and - the one that matters - the
    project's own ``load_companies`` reads it. A workbook the dashboard can
    open but the scraper cannot is the failure this exists to prevent.
    """
    try:
        book = load_workbook(candidate)
    except Exception as exc:
        raise DashboardError(f"The updated workbook did not reopen cleanly: {exc}") from exc
    try:
        sheet = book[book.sheetnames[0]]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if headers != list(expected_headers):
            raise DashboardError(
                "The updated workbook's header row changed; the write was abandoned."
            )
        if sheet.max_row != expected_rows:
            raise DashboardError(
                f"The updated workbook has {sheet.max_row} rows, expected {expected_rows}; "
                "the write was abandoned."
            )
    finally:
        book.close()

    try:
        loaded = load_companies(settings, excel_path=candidate)
    except Exception as exc:
        raise DashboardError(
            f"The scraper's own workbook loader could not read the update: {exc}"
        ) from exc
    if loaded.empty:
        raise DashboardError(
            "The updated workbook contains no companies; the write was abandoned."
        )


def _guard_write(settings: Settings, path: Path) -> None:
    if is_run_active(settings):
        raise WorkbookBusy(
            "A scraper run is in progress. The workbook is read-only until it finishes - "
            "the run reads it, and writes discovered ATS URLs back into it."
        )
    if workbook_is_open_in_excel(path):
        raise WorkbookBusy(
            f"{path.name} is open in Excel: {excel_lock_file(path).name} is present beside it, "
            "or the file is write-locked. Close it in Excel and try again - nothing was "
            f"changed. If Excel is already closed, {excel_lock_file(path).name} was left behind "
            "by a crash and can be deleted."
        )


def _save_atomically(
    book, path: Path, settings: Settings, expected_headers: list, expected_rows: int
) -> Path:
    """Write to a temp file, validate it, back the original up, then swap.

    ``os.replace`` is atomic on both platforms, so a reader never sees a
    half-written workbook: either the old file or the new one, never neither.
    The backup is a single fixed name, overwritten - one recoverable copy, not
    a growing pile beside the ``.bak-*`` files the pipeline already leaves.
    """
    temp = path.with_name(path.name + ".dashboard-tmp.xlsx")
    backup = workbook_backup_path(path)
    try:
        book.save(temp)
    except Exception as exc:
        temp.unlink(missing_ok=True)
        raise DashboardError(f"Could not write the updated workbook: {exc}") from exc

    try:
        _validate_written_workbook(temp, settings, expected_headers, expected_rows)
        shutil.copy2(path, backup)
        os.replace(temp, path)
    except DashboardError:
        temp.unlink(missing_ok=True)
        raise
    except OSError as exc:
        temp.unlink(missing_ok=True)
        raise WorkbookBusy(
            f"Could not replace {path.name}: {exc}. It may have been opened while the update "
            "was being prepared - nothing was changed."
        ) from exc
    return backup


def _header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        if cell.value is not None
    }


@dataclass
class WriteResult:
    """What a successful workbook write did."""

    company: str
    row: int
    backup: Path
    warnings: list[str] = field(default_factory=list)


def add_company(
    name: str,
    ats_url: str = "",
    live_url: str = "",
    *,
    settings: Settings | None = None,
    allow_duplicate_url: bool = False,
) -> WriteResult:
    """Append one company to the workbook, safely.

    Every guard the workbook needs, in order: no run in progress, not open in
    Excel, an exclusive lock, validation of the inputs, duplicate detection,
    then write-temp -> validate -> back up -> atomic swap. Unrelated rows,
    columns, sheets and formatting are untouched because the file is edited
    with openpyxl in place rather than rewritten from a DataFrame.
    """
    cfg = settings or load_settings()
    path = workbook_path(cfg)
    columns = workbook_columns(cfg)

    company = normalize_name(name)
    if not company:
        raise DashboardError("Company name is required.")
    ats_url = str(ats_url or "").strip()
    live_url = str(live_url or "").strip()
    for candidate in (ats_url, live_url):
        problem = validate_url(candidate)
        if problem:
            raise DashboardError(problem)

    _guard_write(cfg, path)

    with _WorkbookLock(path):
        _guard_write(cfg, path)  # re-checked under the lock, not just before it
        existing = read_workbook(cfg)
        duplicates = check_duplicates(existing, columns, company, ats_url, live_url)
        if duplicates.has_name_conflict:
            raise DashboardError(
                f"{duplicates.name_conflict!r} is already in the workbook. "
                "Company names are the scraper's key for a row - edit that row instead."
            )
        warnings: list[str] = []
        if duplicates.has_url_conflict:
            described = "; ".join(
                f"{owner} already uses this URL in {column!r}"
                for owner, column, _ in duplicates.url_conflicts
            )
            if not allow_duplicate_url:
                raise DashboardError(
                    f"Duplicate URL: {described}. Nothing was written. Tick "
                    "'Add anyway' if two companies really do share a job board."
                )
            warnings.append(f"Added despite a duplicate URL: {described}.")

        book = load_workbook(path)
        try:
            sheet = book[book.sheetnames[0]]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            header_map = _header_map(sheet)
            missing = [heading for heading in columns.values() if heading not in header_map]
            if missing:
                raise DashboardError(
                    f"Workbook {path.name} is missing expected column(s): {missing}."
                )

            row = sheet.max_row + 1
            sheet.cell(
                row=row, column=header_map[columns["company"]]
            ).value = escape_formula(company)
            if ats_url:
                sheet.cell(
                    row=row, column=header_map[columns["ats_url"]]
                ).value = escape_formula(ats_url)
            if live_url:
                sheet.cell(
                    row=row, column=header_map[columns["live_jobs_url"]]
                ).value = escape_formula(live_url)

            backup = _save_atomically(book, path, cfg, headers, row)
        finally:
            book.close()

    return WriteResult(company=company, row=row, backup=backup, warnings=warnings)


def update_company(
    name: str,
    ats_url: str = "",
    live_url: str = "",
    *,
    settings: Settings | None = None,
    allow_duplicate_url: bool = False,
) -> WriteResult:
    """Replace one existing company's two URL cells. Nothing else is touched.

    Editing is limited to the URL columns on purpose. The company name is the
    key the run report, the SQLite job ids and the workbook write-back all
    match on, so renaming a row from here would orphan its history; and a
    general cell editor over a workbook the pipeline also writes is exactly the
    complexity this dashboard is meant to avoid.
    """
    cfg = settings or load_settings()
    path = workbook_path(cfg)
    columns = workbook_columns(cfg)

    company = normalize_name(name)
    if not company:
        raise DashboardError("Company name is required.")
    ats_url = str(ats_url or "").strip()
    live_url = str(live_url or "").strip()
    for candidate in (ats_url, live_url):
        problem = validate_url(candidate)
        if problem:
            raise DashboardError(problem)

    _guard_write(cfg, path)

    with _WorkbookLock(path):
        _guard_write(cfg, path)
        existing = read_workbook(cfg)
        duplicates = check_duplicates(
            existing, columns, company, ats_url, live_url, ignore_row_name=company
        )
        warnings: list[str] = []
        if duplicates.has_url_conflict:
            described = "; ".join(
                f"{owner} already uses this URL in {column!r}"
                for owner, column, _ in duplicates.url_conflicts
            )
            if not allow_duplicate_url:
                raise DashboardError(f"Duplicate URL: {described}. Nothing was written.")
            warnings.append(f"Saved despite a duplicate URL: {described}.")

        book = load_workbook(path)
        try:
            sheet = book[book.sheetnames[0]]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            header_map = _header_map(sheet)
            company_column = header_map.get(columns["company"])
            if not company_column:
                raise DashboardError(
                    f"Workbook {path.name} has no {columns['company']!r} column."
                )

            target = None
            for index in range(2, sheet.max_row + 1):
                cell_name = normalize_name(sheet.cell(row=index, column=company_column).value)
                if cell_name.lower() == company.lower():
                    target = index
                    break
            if target is None:
                raise DashboardError(f"{company!r} is not in the workbook.")

            sheet.cell(row=target, column=header_map[columns["ats_url"]]).value = (
                escape_formula(ats_url) if ats_url else None
            )
            sheet.cell(row=target, column=header_map[columns["live_jobs_url"]]).value = (
                escape_formula(live_url) if live_url else None
            )

            backup = _save_atomically(book, path, cfg, headers, sheet.max_row)
        finally:
            book.close()

    return WriteResult(company=company, row=target, backup=backup, warnings=warnings)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "nat"} else text


def companies_for_editing(settings: Settings | None = None) -> list[dict[str, str]]:
    """``{name, ats_url, live_jobs_url}`` per row, for the edit form."""
    cfg = settings or load_settings()
    columns = workbook_columns(cfg)
    frame = read_workbook(cfg)
    rows: list[dict[str, str]] = []
    for _, row in frame.iterrows():
        company = normalize_name(row.get(columns["company"]))
        if not company:
            continue
        rows.append({
            "name": company,
            "ats_url": _cell_text(row.get(columns["ats_url"])),
            "live_jobs_url": _cell_text(row.get(columns["live_jobs_url"])),
        })
    return rows


def filter_jobs(
    jobs: pd.DataFrame,
    *,
    companies: Iterable[str] = (),
    title: str = "",
    location: str = "",
    statuses: Iterable[str] = (),
    posted_from: Any = None,
    posted_to: Any = None,
) -> pd.DataFrame:
    """The Run Scraper tab's filters, as one pure function over the export.

    Rows with no parseable ``date_posted`` survive a date filter deliberately:
    the pipeline keeps and flags undated jobs rather than discarding them (see
    ``date_filter_status``), and a filter that silently dropped them here would
    undo that.
    """
    if jobs.empty:
        return jobs
    frame = jobs
    companies = [company for company in companies if company]
    if companies and "company" in frame.columns:
        frame = frame[frame["company"].isin(companies)]
    if title and "title" in frame.columns:
        frame = frame[
            frame["title"].astype(str).str.contains(title, case=False, na=False, regex=False)
        ]
    if location and "location" in frame.columns:
        frame = frame[
            frame["location"].astype(str).str.contains(
                location, case=False, na=False, regex=False
            )
        ]
    statuses = [status.lower() for status in statuses if status]
    if statuses and "change_status" in frame.columns:
        frame = frame[frame["change_status"].astype(str).str.lower().isin(statuses)]
    if (posted_from or posted_to) and "date_posted" in frame.columns:
        posted = pd.to_datetime(frame["date_posted"], errors="coerce", utc=True, format="mixed")
        keep = pd.Series(True, index=frame.index)
        if posted_from:
            keep &= posted.isna() | (posted >= pd.Timestamp(posted_from, tz="UTC"))
        if posted_to:
            end = pd.Timestamp(posted_to, tz="UTC") + timedelta(days=1)
            keep &= posted.isna() | (posted < end)
        frame = frame[keep]
    return frame.reset_index(drop=True)

"""Write newly discovered ATS URLs back into the input workbook.

When the Playwright search fallback (``browser/playwright_scraper.py``)
uncovers a real ATS behind a branded career page, that discovery only helps
*this* run's diagnostics unless it's persisted somewhere the router will look
next time. This module fills the blank ``ATS URL`` cells in
``config/companies.xlsx`` in place - never overwriting a value already there,
so manual edits to the workbook are always safe - after backing the file up.
"""

from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Safe only because ats/discovery.py keeps its Playwright imports lazy (inside
# functions) - importing this module at module load time must never drag the
# browser stack into a plain "write the spreadsheet" code path.
from ats.discovery import NOT_FOUND, Discovery
from logger import get_logger

log = get_logger("export_ats_urls")


def _blank(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    return text in {"", "nan", "none", "n/a", "na", "-", "null"}


def _backup_path(companies_path: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return companies_path.with_name(f"{companies_path.name}.bak-{timestamp}")


def write_discovered_urls(
    companies_path: Path | str,
    discoveries: dict[str, str],
    *,
    company_column: str = "Company",
    ats_url_column: str = "ATS URL",
) -> dict[str, Any]:
    """Fill blank ATS URL cells for companies discovered this run.

    Args:
        companies_path: path to the workbook to update in place.
        discoveries: ``{company_name: discovered_ats_url}``.
        company_column / ats_url_column: header names to match, matching the
          columns used throughout the pipeline's ``config/settings.yaml``.

    Returns:
        ``{"updated": int, "backup_path": Path | None}``. ``updated`` counts
        cells actually written - a company with no blank ATS URL cell (e.g.
        already resolved by a prior run) is not double-counted.

    Never raises for a missing/malformed workbook path issue that isn't
    fixable here; callers treat this as a best-effort convenience step, not a
    required part of the run.
    """
    path = Path(companies_path)
    if not discoveries or not path.exists():
        return {"updated": 0, "backup_path": None}

    try:
        workbook = load_workbook(path)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {str(cell.value).strip(): cell.column for cell in header_row if cell.value}

        company_col = headers.get(company_column)
        ats_col = headers.get(ats_url_column)
        if not company_col or not ats_col:
            log.warning(
                "Workbook %s missing expected columns (%s, %s); skipping write-back",
                path.name, company_column, ats_url_column,
            )
            return {"updated": 0, "backup_path": None}

        updated = 0
        for row in sheet.iter_rows(min_row=2):
            company_cell = row[company_col - 1]
            ats_cell = row[ats_col - 1]

            company_name = str(company_cell.value).strip() if company_cell.value else ""
            if not company_name or company_name not in discoveries:
                continue
            if not _blank(ats_cell.value):
                continue

            ats_cell.value = discoveries[company_name]
            updated += 1

        if updated == 0:
            return {"updated": 0, "backup_path": None}

        backup_path = _backup_path(path)
        shutil.copy2(path, backup_path)
        workbook.save(path)
        log.info("Wrote %s discovered ATS URL(s) to %s (backup: %s)",
                  updated, path.name, backup_path.name)
        return {"updated": updated, "backup_path": backup_path}

    except Exception as exc:
        log.warning("Could not write discovered ATS URLs back to %s: %s", path, exc)
        return {"updated": 0, "backup_path": None}


def write_repaired_urls(
    companies_path: Path | str,
    repairs: dict[str, tuple[str, str, str]],
    *,
    company_column: str = "Company",
    ats_url_column: str = "ATS URL",
    live_jobs_url_column: str = "Live Jobs Page (if ATS URL unavailable)",
) -> dict[str, Any]:
    """Overwrite a dead URL with the live one ``url_repair.py`` found.

    Unlike :func:`write_discovered_urls`, this **does** overwrite a non-blank
    cell - but only the exact dead value the repair replaced this run, and
    only after that replacement actually returned jobs. ``repairs`` maps
    ``company -> (source, raw_url, repaired_url)`` where ``source`` is
    ``ats.router.SOURCE_ATS_URL`` or ``SOURCE_LIVE_PAGE`` (which column the
    dead URL came from) and ``raw_url`` is the exact dead value read from the
    workbook at plan time. A row whose cell no longer matches ``raw_url`` is
    skipped rather than overwritten - something else (a manual edit, a prior
    write-back this same run) already changed it, so blindly clobbering it
    would destroy that instead of a confirmed-dead value.

    Returns ``{"updated": int, "backup_path": Path | None}``. Best-effort,
    like the other write-back helpers: a failure here never fails the run.
    """
    path = Path(companies_path)
    if not repairs or not path.exists():
        return {"updated": 0, "backup_path": None}

    from ats.router import SOURCE_ATS_URL  # local import: avoid a cycle at module load

    try:
        workbook = load_workbook(path)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {str(cell.value).strip(): cell.column for cell in header_row if cell.value}

        company_col = headers.get(company_column)
        ats_col = headers.get(ats_url_column)
        live_col = headers.get(live_jobs_url_column)
        if not company_col:
            log.warning("Workbook %s missing %s column; skipping repair write-back",
                        path.name, company_column)
            return {"updated": 0, "backup_path": None}

        updated = 0
        for row in sheet.iter_rows(min_row=2):
            company_cell = row[company_col - 1]
            company_name = str(company_cell.value).strip() if company_cell.value else ""
            if not company_name or company_name not in repairs:
                continue

            source, raw_url, repaired_url = repairs[company_name]
            target_col = ats_col if source == SOURCE_ATS_URL else live_col
            if not target_col:
                continue

            cell = row[target_col - 1]
            current = str(cell.value).strip() if cell.value else ""
            if current != raw_url.strip():
                log.debug(
                    "%s: workbook cell no longer matches the dead URL this run repaired "
                    "(expected %r, found %r); leaving it alone",
                    company_name, raw_url, current,
                )
                continue

            cell.value = repaired_url
            updated += 1

        if updated == 0:
            return {"updated": 0, "backup_path": None}

        backup_path = _backup_path(path)
        shutil.copy2(path, backup_path)
        workbook.save(path)
        log.info("Wrote %s repaired URL(s) to %s (backup: %s)",
                  updated, path.name, backup_path.name)
        return {"updated": updated, "backup_path": backup_path}

    except Exception as exc:
        log.warning("Could not write repaired URLs back to %s: %s", path, exc)
        return {"updated": 0, "backup_path": None}


def write_run_status(
    companies_path: Path | str,
    counts: dict[str, int],
    *,
    company_column: str = "Company",
    status_column: str = "Data Retrieved",
    count_column: str = "Jobs Found",
) -> dict[str, Any]:
    """Record each company's retrieval outcome from the latest full run.

    Writes two columns, both derived from the same job count so they can
    never disagree: ``Data Retrieved`` (TRUE/FALSE) and ``Jobs Found`` (the
    number). The count is the more useful of the two - 2 jobs and 2,500 jobs
    both read as TRUE - but the boolean is what you filter on.

    Unlike :func:`write_discovered_urls`, this **overwrites** the cells every
    run: it reports the outcome of the most recent run, not a value the user
    curated, so a stale TRUE from three runs ago would be worse than useless.
    Either column is created if the workbook does not already have it.

    Args:
        companies_path: path to the workbook to update in place.
        counts: ``{company_name: jobs_collected_this_run}``.
        company_column / status_column / count_column: header names to match.

    Returns:
        ``{"updated": int, "backup_path": Path | None}``.

    Best-effort, like the ATS write-back: a failure here never fails the run.
    """
    path = Path(companies_path)
    if not counts or not path.exists():
        return {"updated": 0, "backup_path": None}

    try:
        workbook = load_workbook(path)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {str(cell.value).strip(): cell.column for cell in header_row if cell.value}

        company_col = headers.get(company_column)
        if not company_col:
            log.warning("Workbook %s missing %s column; skipping status write-back",
                        path.name, company_column)
            return {"updated": 0, "backup_path": None}

        status_col = headers.get(status_column)
        if not status_col:
            status_col = sheet.max_column + 1
            sheet.cell(row=1, column=status_col, value=status_column)

        count_col = headers.get(count_column)
        if not count_col:
            count_col = sheet.max_column + 1
            sheet.cell(row=1, column=count_col, value=count_column)

        updated = 0
        for row in sheet.iter_rows(min_row=2):
            company_cell = row[company_col - 1]
            company_name = str(company_cell.value).strip() if company_cell.value else ""
            if not company_name or company_name not in counts:
                continue
            found = int(counts[company_name])
            sheet.cell(
                row=company_cell.row,
                column=status_col,
                value="TRUE" if found else "FALSE",
            )
            sheet.cell(row=company_cell.row, column=count_col, value=found)
            updated += 1

        if updated == 0:
            return {"updated": 0, "backup_path": None}

        backup_path = _backup_path(path)
        shutil.copy2(path, backup_path)
        workbook.save(path)
        log.info("Wrote retrieval status + job counts for %s company row(s) to %s (backup: %s)",
                 updated, path.name, backup_path.name)
        return {"updated": updated, "backup_path": backup_path}

    except Exception as exc:
        log.warning("Could not write run status back to %s: %s", path, exc)
        return {"updated": 0, "backup_path": None}


def write_suggestions(
    companies_path: Path | str,
    discoveries: list[Discovery],
    *,
    apply: bool = False,
    company_column: str = "Company",
    ats_url_column: str = "ATS URL",
    jobs_page_column: str = "Live Jobs Page (if ATS URL unavailable)",
    status_column: str = "Data Retrieved",
) -> dict[str, Any]:
    """Record discovery results without destroying hand-curated values.

    Suggestions go to three new columns by default. The workbook's own
    ``ATS URL`` / ``Live Jobs Page`` values were curated by hand, and silently
    overwriting them would be hostile - a wrong overwrite is far harder to
    notice than an extra column.

    Two exceptions write the real columns:

    * a **blank** ``ATS URL`` cell is filled with a verified URL, matching
      :func:`write_discovered_urls`' established behaviour;
    * ``apply=True`` promotes suggestions, but only for rows whose
      ``Data Retrieved`` is ``FALSE`` - a value already failing cannot be made
      worse by a verified one.

    Returns ``{"updated": int, "applied": int, "backup_path": Path | None}``.
    """
    path = Path(companies_path)
    if not discoveries or not path.exists():
        return {"updated": 0, "applied": 0, "backup_path": None}

    by_company = {d.company: d for d in discoveries}

    try:
        workbook = load_workbook(path)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = {str(cell.value).strip(): cell.column for cell in header_row if cell.value}

        company_col = headers.get(company_column)
        if not company_col:
            log.warning("Workbook %s missing %s column; skipping suggestions",
                        path.name, company_column)
            return {"updated": 0, "applied": 0, "backup_path": None}

        def _column(name: str) -> int:
            existing = headers.get(name)
            if existing:
                return existing
            index = sheet.max_column + 1
            sheet.cell(row=1, column=index, value=name)
            headers[name] = index
            return index

        suggested_ats_col = _column("Suggested ATS URL")
        suggested_page_col = _column("Suggested Jobs Page")
        notes_col = _column("Discovery Notes")

        ats_col = headers.get(ats_url_column)
        page_col = headers.get(jobs_page_column)
        status_col = headers.get(status_column)

        updated = applied = 0
        for row in sheet.iter_rows(min_row=2):
            company_cell = row[company_col - 1]
            name = str(company_cell.value).strip() if company_cell.value else ""
            found = by_company.get(name)
            if not name or found is None:
                continue

            line = company_cell.row
            sheet.cell(row=line, column=suggested_ats_col,
                       value=found.ats_url or NOT_FOUND)
            sheet.cell(row=line, column=suggested_page_col,
                       value=found.jobs_page or NOT_FOUND)
            sheet.cell(row=line, column=notes_col,
                       value=f"{found.method}: {found.note}" if found.note else found.method)
            updated += 1

            if not found.jobs_found:
                continue

            # Exception 1: fill a blank ATS URL cell.
            if found.ats_url and ats_col and _blank(row[ats_col - 1].value):
                row[ats_col - 1].value = found.ats_url
                applied += 1
                continue

            if not apply:
                continue

            # Exception 2: --apply, restricted to rows already failing.
            failing = status_col and str(row[status_col - 1].value).strip().upper() == "FALSE"
            if not failing:
                continue
            if found.ats_url and ats_col:
                row[ats_col - 1].value = found.ats_url
                applied += 1
            elif found.jobs_page and page_col:
                row[page_col - 1].value = found.jobs_page
                applied += 1

        backup_path = _backup_path(path)
        shutil.copy2(path, backup_path)
        workbook.save(path)
        log.info("Wrote %s suggestion row(s), applied %s, to %s (backup: %s)",
                 updated, applied, path.name, backup_path.name)
        return {"updated": updated, "applied": applied, "backup_path": backup_path}

    except Exception as exc:
        log.warning("Could not write suggestions to %s: %s", path, exc)
        return {"updated": 0, "applied": 0, "backup_path": None}

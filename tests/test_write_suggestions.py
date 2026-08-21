import openpyxl
import pytest

from ats.discovery import Discovery
from export_ats_urls import write_suggestions

HEADERS = ["Company", "ATS URL", "Live Jobs Page (if ATS URL unavailable)",
           "Data Retrieved", "Jobs Found"]


@pytest.fixture
def workbook(tmp_path):
    path = tmp_path / "companies.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["Nokia", None, "https://www.nokia.com/about-us/careers/", "FALSE", 0])
    ws.append(["Infosys", "https://www.infosys.com/careers/",
               "https://www.infosys.com/careers/", "FALSE", 0])
    ws.append(["Capital One", "https://capitalone.wd12.myworkdayjobs.com/Capital_One",
               None, "TRUE", 500])
    wb.save(path)
    return path


def _read(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows[record["Company"]] = record
    wb.close()
    return headers, rows


def test_creates_the_three_suggestion_columns(workbook):
    write_suggestions(workbook, [Discovery(company="Nokia", ats_url="https://fa-x.oraclecloud.com",
                                           provider="taleo", jobs_found=575,
                                           method="http", note="taleo API returned 575 jobs")])
    headers, _ = _read(workbook)
    assert "Suggested ATS URL" in headers
    assert "Suggested Jobs Page" in headers
    assert "Discovery Notes" in headers


def test_verified_finding_is_suggested(workbook):
    write_suggestions(workbook, [Discovery(company="Nokia", ats_url="https://fa-x.oraclecloud.com",
                                           provider="taleo", jobs_found=575,
                                           method="http", note="taleo API returned 575 jobs")])
    _, rows = _read(workbook)
    assert rows["Nokia"]["Suggested ATS URL"] == "https://fa-x.oraclecloud.com"
    assert "575" in rows["Nokia"]["Discovery Notes"]


def test_nothing_found_is_recorded_as_not_found(workbook):
    write_suggestions(workbook, [Discovery(company="Infosys", note="nothing found in page HTML")])
    _, rows = _read(workbook)
    assert rows["Infosys"]["Suggested ATS URL"] == "NOT FOUND"
    assert rows["Infosys"]["Suggested Jobs Page"] == "NOT FOUND"


def test_curated_values_are_never_overwritten_without_apply(workbook):
    write_suggestions(workbook, [Discovery(company="Infosys", ats_url="https://boards.greenhouse.io/infy",
                                           provider="greenhouse", jobs_found=12,
                                           method="http", note="greenhouse API returned 12 jobs")])
    _, rows = _read(workbook)
    assert rows["Infosys"]["ATS URL"] == "https://www.infosys.com/careers/"


def test_blank_ats_cell_is_filled_directly(workbook):
    write_suggestions(workbook, [Discovery(company="Nokia", ats_url="https://fa-x.oraclecloud.com",
                                           provider="taleo", jobs_found=575,
                                           method="http", note="ok")])
    _, rows = _read(workbook)
    assert rows["Nokia"]["ATS URL"] == "https://fa-x.oraclecloud.com"


def test_apply_overwrites_only_failing_rows(workbook):
    write_suggestions(
        workbook,
        [
            Discovery(company="Infosys", ats_url="https://boards.greenhouse.io/infy",
                      provider="greenhouse", jobs_found=12, method="http", note="ok"),
            Discovery(company="Capital One", ats_url="https://example.com/wrong",
                      provider="workday", jobs_found=1, method="http", note="ok"),
        ],
        apply=True,
    )
    _, rows = _read(workbook)
    assert rows["Infosys"]["ATS URL"] == "https://boards.greenhouse.io/infy"
    # Capital One is Data Retrieved = TRUE, so --apply must leave it alone.
    assert rows["Capital One"]["ATS URL"] == "https://capitalone.wd12.myworkdayjobs.com/Capital_One"


def test_a_backup_is_created(workbook):
    result = write_suggestions(workbook, [Discovery(company="Nokia", note="nothing")])
    assert result["backup_path"] is not None
    assert result["backup_path"].exists()

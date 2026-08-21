"""Offline tests for write_repaired_urls.

Regression coverage for the gap where ats/url_repair.py's dead-URL fix was
only ever applied in memory for the current run: NTT DATA's dead
``careers.nttdata.com`` was "repaired" to a live URL and scraped
successfully, but the workbook still held the dead value every subsequent
run, paying the repair cost again each time.
"""

import openpyxl
import pytest

from ats.router import SOURCE_ATS_URL, SOURCE_LIVE_PAGE
from export_ats_urls import write_repaired_urls

HEADERS = ["Company", "ATS URL", "Live Jobs Page (if ATS URL unavailable)"]


@pytest.fixture
def workbook(tmp_path):
    path = tmp_path / "companies.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["NTT DATA", None, "https://careers.nttdata.com/"])
    ws.append(["Primoris Services", "https://careers.prim.com/", None])
    ws.append(["Untouched Co", "https://untouched.example/jobs", None])
    wb.save(path)
    return path


def _read(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        rows[record["Company"]] = record
    wb.close()
    return rows


def test_overwrites_the_dead_live_jobs_url(workbook):
    repairs = {
        "NTT DATA": (
            SOURCE_LIVE_PAGE,
            "https://careers.nttdata.com/",
            "https://www.nttdata.com/en-us/careers",
        ),
    }
    result = write_repaired_urls(workbook, repairs)
    assert result["updated"] == 1

    rows = _read(workbook)
    assert rows["NTT DATA"]["Live Jobs Page (if ATS URL unavailable)"] == \
        "https://www.nttdata.com/en-us/careers"


def test_overwrites_the_dead_ats_url(workbook):
    repairs = {
        "Primoris Services": (
            SOURCE_ATS_URL,
            "https://careers.prim.com/",
            "https://www.prim.com/careers",
        ),
    }
    result = write_repaired_urls(workbook, repairs)
    assert result["updated"] == 1

    rows = _read(workbook)
    assert rows["Primoris Services"]["ATS URL"] == "https://www.prim.com/careers"


def test_skips_a_cell_that_no_longer_matches_the_dead_value(workbook):
    """If the workbook cell isn't the exact dead value repair replaced, don't touch it."""
    repairs = {
        "Untouched Co": (
            SOURCE_ATS_URL,
            "https://this-was-never-the-real-value.example/",
            "https://sneaky-overwrite.example/",
        ),
    }
    result = write_repaired_urls(workbook, repairs)
    assert result["updated"] == 0

    rows = _read(workbook)
    assert rows["Untouched Co"]["ATS URL"] == "https://untouched.example/jobs"


def test_no_repairs_is_a_noop(workbook):
    result = write_repaired_urls(workbook, {})
    assert result == {"updated": 0, "backup_path": None}

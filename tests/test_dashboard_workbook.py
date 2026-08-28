"""Writing ``companies.xlsx`` from the dashboard, safely.

The workbook is the pipeline's only input and one it also writes back to, so
every test here is about not damaging it: the schema survives, unrelated rows
and columns survive, a formula never lands in a cell, a write is atomic, and
the project's own loader can still read the result.
"""

from __future__ import annotations

import shutil

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from dashboard import services
from pipeline import load_companies
from settings import Settings

HEADERS = [
    "Company", "ATS URL", "Live Jobs Page (if ATS URL unavailable)",
    "Data Retrieved", "Jobs Found", "Suggested ATS URL", "Suggested Jobs Page",
    "Discovery Notes",
]

ROWS = [
    ["Capital One", "https://capitalone.wd12.myworkdayjobs.com/en-US/Capital_One/", None,
     "TRUE", 1851, None, None, None],
    ["Boingo Wireless", "https://boards.greenhouse.io/boingo", None, "TRUE", 6, None, None,
     "verified 2026-08-01"],
    ["Goldman Sachs", "https://higher.gs.com/results", "https://www.goldmansachs.com/careers/",
     "TRUE", 820, None, None, None],
]


def build_workbook(path):
    """A workbook shaped like the real one: named sheet, table, column widths."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Company Job Sources"
    sheet.append(HEADERS)
    for row in ROWS:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 34.0
    sheet.column_dimensions["B"].width = 52.0
    table = Table(displayName="CompanyJobSourcesTable", ref="A1:C4")
    table.tableStyleInfo = TableStyleInfo(showRowStripes=True)
    sheet.add_table(table)
    book.save(path)
    book.close()


@pytest.fixture()
def cfg(tmp_path) -> Settings:
    workbook = tmp_path / "companies.xlsx"
    build_workbook(workbook)
    (tmp_path / "output").mkdir()
    return Settings(
        {
            "input_excel": str(workbook),
            "input_sheet": None,
            "columns": {
                "company": "Company",
                "ats_url": "ATS URL",
                "live_jobs_url": "Live Jobs Page (if ATS URL unavailable)",
            },
            "output": {"directory": str(tmp_path / "output")},
            "logging": {"file": str(tmp_path / "logs" / "scraper.log")},
        },
        tmp_path / "settings.yaml",
    )


# ---------------------------------------------------------------------------
# Insertion
# ---------------------------------------------------------------------------

def test_a_company_is_appended_and_the_projects_own_loader_reads_it(cfg):
    result = services.add_company(
        "Acme Data", "https://boards.greenhouse.io/acme", settings=cfg
    )

    assert result.company == "Acme Data"
    assert result.row == 5

    loaded = load_companies(cfg)
    assert "Acme Data" in list(loaded["company"])
    assert loaded.loc[
        loaded["company"] == "Acme Data", "ats_url"
    ].iloc[0] == "https://boards.greenhouse.io/acme"
    assert len(loaded) == 4


def test_a_company_with_no_urls_is_allowed(cfg):
    """Blank ATS URL is the normal case - the pipeline discovers and fills it."""
    services.add_company("Unknown ATS Co", settings=cfg)
    loaded = load_companies(cfg)
    row = loaded[loaded["company"] == "Unknown ATS Co"].iloc[0]
    assert row["company"] == "Unknown ATS Co"


def test_an_empty_name_is_refused(cfg):
    for name in ("", "   ", None):
        with pytest.raises(services.DashboardError, match="required"):
            services.add_company(name, settings=cfg)


def test_names_are_normalised_the_way_the_loader_normalises_them(cfg):
    services.add_company("  Padded Name  ", settings=cfg)
    assert "Padded Name" in list(load_companies(cfg)["company"])


@pytest.mark.parametrize("bad", [
    "not a url at all",
    "ftp://files.example.com/jobs",
    "javascript:alert(1)",
    "https://exa mple.com/jobs",
])
def test_an_unusable_url_is_refused_before_anything_is_written(cfg, bad):
    with pytest.raises(services.DashboardError):
        services.add_company("Bad URL Co", bad, settings=cfg)
    assert len(load_companies(cfg)) == 3


def test_a_bare_hostname_is_accepted_as_https(cfg):
    assert services.validate_url("boards.greenhouse.io/acme") is None
    assert services.validate_url("") is None


# ---------------------------------------------------------------------------
# Duplicates
# ---------------------------------------------------------------------------

def test_a_duplicate_company_name_is_refused_case_insensitively(cfg):
    with pytest.raises(services.DashboardError, match="already in the workbook"):
        services.add_company("  capital ONE ", settings=cfg)
    assert len(load_companies(cfg)) == 3


def test_a_duplicate_url_is_refused_until_it_is_explicitly_allowed(cfg):
    # Trailing slash, casing and a www. prefix must not hide the duplicate.
    duplicate = "https://WWW.Boards.Greenhouse.io/boingo/"
    with pytest.raises(services.DashboardError, match="Duplicate URL"):
        services.add_company("Boingo Clone", duplicate, settings=cfg)
    assert len(load_companies(cfg)) == 3

    result = services.add_company(
        "Boingo Clone", duplicate, settings=cfg, allow_duplicate_url=True
    )
    assert result.warnings and "duplicate URL" in result.warnings[0]
    assert len(load_companies(cfg)) == 4


def test_a_duplicate_is_caught_across_both_url_columns(cfg):
    """Goldman's careers page is in Live Jobs Page; proposing it as an ATS URL
    is still the same site."""
    with pytest.raises(services.DashboardError, match="Duplicate URL"):
        services.add_company(
            "Goldman Clone", "https://www.goldmansachs.com/careers", settings=cfg
        )


def test_duplicate_detection_reports_which_row_owns_the_url(cfg):
    frame = services.read_workbook(cfg)
    check = services.check_duplicates(
        frame, services.workbook_columns(cfg), "New Co",
        ats_url="https://boards.greenhouse.io/boingo",
    )
    assert check.has_url_conflict
    assert check.url_conflicts[0][0] == "Boingo Wireless"
    assert check.has_name_conflict is False


def test_url_normalisation_collapses_the_noise():
    same = services.normalize_url("https://WWW.Example.com/Jobs/")
    assert same == services.normalize_url("http://example.com/Jobs")
    assert same == services.normalize_url("example.com/Jobs#anchor")
    assert services.normalize_url("") == ""
    assert services.normalize_url(None) == ""
    assert services.normalize_url("nan") == ""


# ---------------------------------------------------------------------------
# Formula injection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("payload", [
    "=HYPERLINK(\"http://evil\",\"click\")",
    "+1+1",
    "-2-2",
    "@SUM(A1:A9)",
])
def test_a_formula_typed_into_the_form_lands_as_text(cfg, payload):
    services.add_company(payload, settings=cfg)

    book = load_workbook(services.workbook_path(cfg))
    value = book["Company Job Sources"].cell(row=5, column=1).value
    book.close()
    assert value.startswith("'")
    assert value[1:] == payload


def test_escaping_leaves_ordinary_text_alone():
    assert services.escape_formula("Capital One") == "Capital One"
    assert services.escape_formula("=cmd") == "'=cmd"
    assert services.escape_formula("\tTabbed") == "'\tTabbed"


# ---------------------------------------------------------------------------
# Schema and formatting preservation
# ---------------------------------------------------------------------------

def test_every_column_sheet_and_unrelated_row_survives_a_write(cfg):
    path = services.workbook_path(cfg)
    services.add_company("Acme Data", "https://boards.greenhouse.io/acme", settings=cfg)

    book = load_workbook(path)
    try:
        assert book.sheetnames == ["Company Job Sources"]
        sheet = book["Company Job Sources"]
        assert [cell.value for cell in sheet[1]] == HEADERS
        # Unrelated rows, including the columns the dashboard never touches.
        assert [sheet.cell(row=3, column=index).value for index in (1, 4, 5, 8)] == [
            "Boingo Wireless", "TRUE", 6, "verified 2026-08-01"
        ]
        # Formatting and the defined table are preserved by editing in place.
        assert sheet.column_dimensions["A"].width == 34.0
        assert "CompanyJobSourcesTable" in sheet.tables
        # The new row only fills the three columns the schema defines for it.
        assert [sheet.cell(row=5, column=index).value for index in range(4, 9)] == [None] * 5
    finally:
        book.close()


def test_a_workbook_missing_a_required_column_is_refused(cfg, tmp_path):
    book = Workbook()
    book.active.append(["Company", "Something Else"])
    book.active.append(["Only Co", "x"])
    book.save(services.workbook_path(cfg))
    book.close()

    with pytest.raises(services.DashboardError, match="missing expected column"):
        services.add_company("New Co", settings=cfg)


# ---------------------------------------------------------------------------
# Atomicity
# ---------------------------------------------------------------------------

def test_a_write_that_fails_validation_leaves_the_original_untouched(cfg, monkeypatch):
    path = services.workbook_path(cfg)
    before = path.read_bytes()

    monkeypatch.setattr(
        services, "_validate_written_workbook",
        lambda *args, **kwargs: (_ for _ in ()).throw(services.DashboardError("nope")),
    )
    with pytest.raises(services.DashboardError, match="nope"):
        services.add_company("Acme Data", settings=cfg)

    assert path.read_bytes() == before
    assert not path.with_name(path.name + ".dashboard-tmp.xlsx").exists()
    assert len(load_companies(cfg)) == 3


def test_a_temp_workbook_the_loader_cannot_read_is_never_promoted(cfg, monkeypatch):
    """The last line of defence: validation drives the *real* loader."""
    path = services.workbook_path(cfg)
    before = path.read_bytes()

    def broken(settings, excel_path=None):
        raise ValueError("Workbook is missing expected column(s)")

    monkeypatch.setattr(services, "load_companies", broken)
    with pytest.raises(services.DashboardError, match="own workbook loader"):
        services.add_company("Acme Data", settings=cfg)
    assert path.read_bytes() == before


def test_exactly_one_backup_is_kept(cfg):
    path = services.workbook_path(cfg)
    services.add_company("First Co", settings=cfg)
    services.add_company("Second Co", settings=cfg)

    backup = services.workbook_backup_path(path)
    assert sorted(p.name for p in path.parent.glob("*.bak*")) == [backup.name]
    # It keeps the .xlsx extension, so recovery is a rename away, not a repair.
    assert backup.suffix == ".xlsx"
    # The single backup is the state immediately before the latest write.
    restored = load_workbook(backup)
    names = [row[0] for row in restored["Company Job Sources"].iter_rows(min_row=2, values_only=True)]
    restored.close()
    assert "First Co" in names and "Second Co" not in names


def test_no_stray_lock_or_temp_file_is_left_behind(cfg):
    path = services.workbook_path(cfg)
    services.add_company("Acme Data", settings=cfg)

    leftovers = [
        p.name for p in path.parent.iterdir()
        if p.name.endswith(services.WORKBOOK_LOCK_SUFFIX) or "dashboard-tmp" in p.name
    ]
    assert leftovers == []


# ---------------------------------------------------------------------------
# Locking
# ---------------------------------------------------------------------------

def test_a_write_is_refused_while_the_scraper_is_running(cfg, monkeypatch):
    monkeypatch.setattr(services, "is_run_active", lambda settings=None: True)

    with pytest.raises(services.WorkbookBusy, match="scraper run is in progress"):
        services.add_company("Acme Data", settings=cfg)
    with pytest.raises(services.WorkbookBusy):
        services.update_company("Capital One", "https://example.com/x", settings=cfg)
    assert len(load_companies(cfg)) == 3


def test_a_workbook_open_in_excel_is_never_touched(cfg):
    path = services.workbook_path(cfg)
    before = path.read_bytes()
    services.excel_lock_file(path).write_bytes(b"excel owner file")

    assert services.workbook_is_open_in_excel(path) is True
    with pytest.raises(services.WorkbookBusy, match="open in Excel"):
        services.add_company("Acme Data", settings=cfg)
    assert path.read_bytes() == before


def test_a_competing_writer_holds_the_lock(cfg):
    path = services.workbook_path(cfg)
    with services._WorkbookLock(path):
        with pytest.raises(services.WorkbookBusy, match="Another workbook write"):
            services.add_company("Acme Data", settings=cfg)
    # The lock is released on exit, so the next write succeeds.
    services.add_company("Acme Data", settings=cfg)
    assert len(load_companies(cfg)) == 4


def test_the_workbook_lock_is_reentrant_across_sequential_writes(cfg):
    services.add_company("One", settings=cfg)
    services.add_company("Two", settings=cfg)
    assert len(load_companies(cfg)) == 5


def test_a_lock_left_by_a_dead_writer_is_broken_after_its_deadline(cfg, monkeypatch):
    path = services.workbook_path(cfg)
    lock = path.with_name(path.name + services.WORKBOOK_LOCK_SUFFIX)
    lock.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(services, "WORKBOOK_LOCK_STALE_SECONDS", -1)

    services.add_company("Acme Data", settings=cfg)
    assert len(load_companies(cfg)) == 4


def test_the_write_lock_times_out_rather_than_hanging(cfg, monkeypatch):
    path = services.workbook_path(cfg)
    lock = path.with_name(path.name + services.WORKBOOK_LOCK_SUFFIX)
    lock.write_text("{}", encoding="utf-8")
    monkeypatch.setattr(services, "WORKBOOK_LOCK_TIMEOUT_SECONDS", 0.2)

    with pytest.raises(services.WorkbookBusy):
        services.add_company("Acme Data", settings=cfg)


# ---------------------------------------------------------------------------
# Editing
# ---------------------------------------------------------------------------

def test_an_existing_company_can_have_its_urls_edited(cfg):
    services.update_company(
        "Capital One",
        "https://capitalone.wd12.myworkdayjobs.com/en-US/Capital_One/",
        "https://www.capitalonecareers.com/",
        settings=cfg,
    )

    loaded = load_companies(cfg)
    row = loaded[loaded["company"] == "Capital One"].iloc[0]
    assert row["live_jobs_url"] == "https://www.capitalonecareers.com/"
    assert len(loaded) == 3  # an edit never adds a row


def test_editing_can_clear_a_url(cfg):
    services.update_company("Capital One", "", "", settings=cfg)
    loaded = load_companies(cfg)
    row = loaded[loaded["company"] == "Capital One"].iloc[0]
    assert str(row["ats_url"]) in ("nan", "None", "")


def test_editing_an_unknown_company_is_refused(cfg):
    with pytest.raises(services.DashboardError, match="not in the workbook"):
        services.update_company("Nobody Inc", "https://example.com/jobs", settings=cfg)


def test_a_row_keeps_its_own_url_without_tripping_duplicate_detection(cfg):
    services.update_company(
        "Boingo Wireless", "https://boards.greenhouse.io/boingo", settings=cfg
    )
    assert len(load_companies(cfg)) == 3


def test_editing_onto_another_companys_url_is_refused(cfg):
    with pytest.raises(services.DashboardError, match="Duplicate URL"):
        services.update_company(
            "Capital One", "https://boards.greenhouse.io/boingo", settings=cfg
        )


def test_the_edit_form_lists_every_company_with_its_current_urls(cfg):
    rows = services.companies_for_editing(cfg)
    assert [row["name"] for row in rows] == ["Capital One", "Boingo Wireless", "Goldman Sachs"]
    assert rows[0]["live_jobs_url"] == ""  # blank cells read as blank, not "nan"
    assert rows[2]["live_jobs_url"] == "https://www.goldmansachs.com/careers/"


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def test_the_table_shows_every_workbook_column(cfg):
    frame = services.read_workbook(cfg)
    assert list(frame.columns) == HEADERS
    assert len(frame) == 3


def test_a_missing_workbook_is_a_clear_message(cfg):
    services.workbook_path(cfg).unlink()
    with pytest.raises(services.DashboardError, match="Workbook not found"):
        services.read_workbook(cfg)


def test_the_real_projects_workbook_columns_are_the_configured_ones(cfg):
    assert services.workbook_columns(cfg) == {
        "company": "Company",
        "ats_url": "ATS URL",
        "live_jobs_url": "Live Jobs Page (if ATS URL unavailable)",
    }


def test_a_copy_of_the_shipped_example_workbook_accepts_a_company(tmp_path):
    """End to end against the workbook the repo actually ships."""
    from settings import PROJECT_ROOT

    source = PROJECT_ROOT / "config" / "companies.example.xlsx"
    if not source.exists():  # pragma: no cover - the example is tracked
        pytest.skip("config/companies.example.xlsx is not present")
    workbook = tmp_path / "companies.xlsx"
    shutil.copy2(source, workbook)

    cfg = Settings(
        {
            "input_excel": str(workbook),
            "input_sheet": None,
            "columns": {
                "company": "Company",
                "ats_url": "ATS URL",
                "live_jobs_url": "Live Jobs Page (if ATS URL unavailable)",
            },
            "output": {"directory": str(tmp_path / "output")},
        },
        tmp_path / "settings.yaml",
    )
    before = len(load_companies(cfg))
    services.add_company("Temporary Test Co", "https://boards.greenhouse.io/tmp", settings=cfg)
    after = load_companies(cfg)

    assert len(after) == before + 1
    assert "Temporary Test Co" in list(after["company"])
